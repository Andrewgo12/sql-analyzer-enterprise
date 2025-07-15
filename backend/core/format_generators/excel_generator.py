"""
Excel Workbook Generator

Generates multi-sheet Excel workbooks with charts and analysis data.
"""

import time
import io
from typing import Dict, Any, List
from .base_generator import BaseFormatGenerator, GenerationContext, GenerationResult

# Optional imports for Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelWorkbookGenerator(BaseFormatGenerator):
    """Generator for Excel workbooks with multiple sheets and charts."""
    
    @property
    def format_name(self) -> str:
        return "Libro Excel"
    
    @property
    def file_extension(self) -> str:
        return ".xlsx"
    
    @property
    def mime_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    @property
    def is_binary(self) -> bool:
        return True
    
    def generate(self, context: GenerationContext) -> GenerationResult:
        """Generate Excel workbook with multiple sheets."""
        start_time = time.time()
        
        try:
            if not OPENPYXL_AVAILABLE:
                raise Exception("openpyxl no está disponible. Instale con: pip install openpyxl")
            
            self.validate_context(context)
            
            # Generate Excel content
            excel_content = self._generate_excel_workbook(context)
            
            generation_time = time.time() - start_time
            
            return self.create_generation_result(
                excel_content, context, generation_time,
                metadata={
                    "sheets_created": 5,
                    "charts_included": 2,
                    "format": "Excel 2007+"
                }
            )
            
        except Exception as e:
            return self.handle_generation_error(e, context)
    
    def _generate_excel_workbook(self, context: GenerationContext) -> bytes:
        """Generate the complete Excel workbook."""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Get template variables
        template_vars = self.get_template_variables(context)
        
        # Create sheets
        self._create_summary_sheet(wb, template_vars)
        self._create_errors_sheet(wb, template_vars)
        self._create_statistics_sheet(wb, template_vars)
        self._create_recommendations_sheet(wb, template_vars)
        self._create_charts_sheet(wb, template_vars)
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_content = buffer.getvalue()
        buffer.close()
        
        return excel_content
    
    def _create_summary_sheet(self, wb: Workbook, template_vars: Dict[str, Any]):
        """Create summary sheet."""
        ws = wb.create_sheet("Resumen", 0)
        summary = template_vars['summary']
        
        # Title
        ws['A1'] = "📊 Reporte de Análisis SQL"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        ws.merge_cells('A1:D1')
        
        # File information
        ws['A3'] = "Información del Archivo"
        ws['A3'].font = Font(bold=True)
        
        file_info = [
            ("Archivo:", template_vars['original_filename']),
            ("Fecha de Análisis:", template_vars['analysis_timestamp'].strftime('%d/%m/%Y %H:%M:%S')),
            ("Sesión:", template_vars['session_id']),
            ("Generador:", template_vars['generator_name'])
        ]
        
        for i, (label, value) in enumerate(file_info, 4):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Analysis summary
        ws['A9'] = "Resumen del Análisis"
        ws['A9'].font = Font(bold=True)
        
        analysis_data = [
            ("Total de Errores:", summary['total_errors']),
            ("Errores Críticos:", summary['critical_errors']),
            ("Errores Altos:", summary['high_errors']),
            ("Errores Medios:", summary['medium_errors']),
            ("Errores Bajos:", summary['low_errors']),
            ("Puntuación de Calidad:", f"{summary['quality_score']}%"),
            ("Líneas Analizadas:", summary['lines_analyzed']),
            ("Correcciones Sugeridas:", summary['fixes_suggested'])
        ]
        
        for i, (label, value) in enumerate(analysis_data, 10):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
            
            # Color code based on severity
            if "Críticos" in label and value > 0:
                ws[f'B{i}'].fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            elif "Puntuación" in label:
                if isinstance(value, str) and value.replace('%', '').isdigit():
                    score = int(value.replace('%', ''))
                    if score < 70:
                        ws[f'B{i}'].fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                    elif score >= 85:
                        ws[f'B{i}'].fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_errors_sheet(self, wb: Workbook, template_vars: Dict[str, Any]):
        """Create errors detail sheet."""
        ws = wb.create_sheet("Errores Detallados")
        
        # Headers
        headers = ["ID", "Severidad", "Categoría", "Título", "Línea", "Columna", "Descripción", "Confianza"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Error data
        for row, error in enumerate(template_vars['errors'], 2):
            ws.cell(row=row, column=1, value=error.get('id', ''))
            ws.cell(row=row, column=2, value=error.get('severity_display', ''))
            ws.cell(row=row, column=3, value=error.get('category', ''))
            ws.cell(row=row, column=4, value=error.get('title', ''))
            ws.cell(row=row, column=5, value=error.get('line', 0))
            ws.cell(row=row, column=6, value=error.get('column', 0))
            ws.cell(row=row, column=7, value=error.get('description', ''))
            ws.cell(row=row, column=8, value=error.get('confidence', 0))
            
            # Color code by severity
            severity = error.get('severity', '').lower()
            if severity == 'critical':
                fill_color = "FFEBEE"
            elif severity == 'high':
                fill_color = "FFF3E0"
            elif severity == 'medium':
                fill_color = "FFFDE7"
            elif severity == 'low':
                fill_color = "E8F5E8"
            else:
                fill_color = "F5F5F5"
            
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_statistics_sheet(self, wb: Workbook, template_vars: Dict[str, Any]):
        """Create statistics sheet."""
        ws = wb.create_sheet("Estadísticas")
        
        # Title
        ws['A1'] = "📈 Estadísticas de Análisis"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:C1')
        
        # Error distribution by severity
        ws['A3'] = "Distribución por Severidad"
        ws['A3'].font = Font(bold=True)
        
        summary = template_vars['summary']
        severity_data = [
            ("Severidad", "Cantidad", "Porcentaje"),
            ("Críticos", summary['critical_errors'], f"{(summary['critical_errors']/max(1,summary['total_errors'])*100):.1f}%"),
            ("Altos", summary['high_errors'], f"{(summary['high_errors']/max(1,summary['total_errors'])*100):.1f}%"),
            ("Medios", summary['medium_errors'], f"{(summary['medium_errors']/max(1,summary['total_errors'])*100):.1f}%"),
            ("Bajos", summary['low_errors'], f"{(summary['low_errors']/max(1,summary['total_errors'])*100):.1f}%")
        ]
        
        for row, (severity, count, percentage) in enumerate(severity_data, 4):
            ws.cell(row=row, column=1, value=severity)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=percentage)
            
            if row == 4:  # Header row
                for col in range(1, 4):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        
        # Error distribution by category
        ws['A10'] = "Distribución por Categoría"
        ws['A10'].font = Font(bold=True)
        
        categories = template_vars['categories']
        category_headers = ["Categoría", "Cantidad"]
        
        for col, header in enumerate(category_headers, 1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        
        for row, (category, count) in enumerate(categories.items(), 12):
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=count)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_recommendations_sheet(self, wb: Workbook, template_vars: Dict[str, Any]):
        """Create recommendations sheet."""
        ws = wb.create_sheet("Recomendaciones")
        
        # Title
        ws['A1'] = "💡 Recomendaciones"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:C1')
        
        summary = template_vars['summary']
        recommendations = []
        
        if summary['critical_errors'] > 0:
            recommendations.append(("🚨 URGENTE", "Corrija los errores críticos antes de ejecutar el código", "ALTA"))
        
        if summary['high_errors'] > 0:
            recommendations.append(("⚠️ IMPORTANTE", "Revise y corrija los errores de alta prioridad", "ALTA"))
        
        if summary['quality_score'] < 70:
            recommendations.append(("📈 MEJORA", "La puntuación de calidad es baja, considere refactorizar", "MEDIA"))
        
        if summary['fixes_suggested'] > 0:
            recommendations.append(("🔧 AUTOMATIZACIÓN", f"{summary['fixes_suggested']} correcciones pueden aplicarse automáticamente", "MEDIA"))
        
        recommendations.extend([
            ("📚 DOCUMENTACIÓN", "Revise la documentación para mejores prácticas", "BAJA"),
            ("🔍 REVISIÓN", "Implemente revisiones de código regulares", "BAJA"),
            ("🧪 TESTING", "Pruebe el código en un entorno de desarrollo", "MEDIA")
        ])
        
        # Headers
        headers = ["Prioridad", "Recomendación", "Importancia"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Recommendations data
        for row, (priority, recommendation, importance) in enumerate(recommendations, 4):
            ws.cell(row=row, column=1, value=priority)
            ws.cell(row=row, column=2, value=recommendation)
            ws.cell(row=row, column=3, value=importance)
            
            # Color code by importance
            if importance == "ALTA":
                fill_color = "FFEBEE"
            elif importance == "MEDIA":
                fill_color = "FFF3E0"
            else:
                fill_color = "E8F5E8"
            
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 15
    
    def _create_charts_sheet(self, wb: Workbook, template_vars: Dict[str, Any]):
        """Create charts sheet."""
        ws = wb.create_sheet("Gráficos")
        
        # Title
        ws['A1'] = "📊 Análisis Visual"
        ws['A1'].font = Font(size=14, bold=True)
        
        summary = template_vars['summary']
        
        # Severity pie chart data
        ws['A3'] = "Distribución por Severidad"
        ws['A3'].font = Font(bold=True)
        
        severity_chart_data = [
            ("Severidad", "Cantidad"),
            ("Críticos", summary['critical_errors']),
            ("Altos", summary['high_errors']),
            ("Medios", summary['medium_errors']),
            ("Bajos", summary['low_errors'])
        ]
        
        for row, (label, value) in enumerate(severity_chart_data, 4):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
        
        # Create pie chart
        pie_chart = PieChart()
        pie_chart.title = "Distribución de Errores por Severidad"
        
        data = Reference(ws, min_col=2, min_row=5, max_row=8, max_col=2)
        labels = Reference(ws, min_col=1, min_row=5, max_row=8, max_col=1)
        
        pie_chart.add_data(data)
        pie_chart.set_categories(labels)
        pie_chart.height = 10
        pie_chart.width = 15
        
        ws.add_chart(pie_chart, "D3")
        
        # Category bar chart data
        categories = template_vars['categories']
        if categories:
            ws['A15'] = "Distribución por Categoría"
            ws['A15'].font = Font(bold=True)
            
            category_data = [("Categoría", "Cantidad")]
            category_data.extend(list(categories.items())[:10])  # Limit to top 10
            
            for row, (label, value) in enumerate(category_data, 16):
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=value)
            
            # Create bar chart
            bar_chart = BarChart()
            bar_chart.title = "Errores por Categoría"
            
            data = Reference(ws, min_col=2, min_row=17, max_row=16+len(categories), max_col=2)
            labels = Reference(ws, min_col=1, min_row=17, max_row=16+len(categories), max_col=1)
            
            bar_chart.add_data(data)
            bar_chart.set_categories(labels)
            bar_chart.height = 10
            bar_chart.width = 15
            
            ws.add_chart(bar_chart, "D15")
